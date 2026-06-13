import io
import datetime
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from apartment_maintenance.models import Flat, Payment


def get_flat_data():
    return list(
        Flat.objects.select_related('wing', 'floor', 'owner')
        .prefetch_related('payments', 'maintenance_dues')
        .order_by('wing__name', 'floor__number', 'room_number')
    )

STATUS_COLORS_PDF = {
    'paid':    colors.HexColor('#dcfce7'),
    'pending': colors.HexColor('#fef9c3'),
    'overdue': colors.HexColor('#fee2e2'),
}


# ── ALL FLATS PDF ──────────────────────────────────────────────────────
class ExportPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        flats = get_flat_data()
        wing          = request.query_params.get('wing')
        status_filter = request.query_params.get('status')
        if wing:
            flats = [f for f in flats if f.wing.name == wing.upper()]
        if status_filter:
            flats = [f for f in flats if f.payment_status == status_filter]

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
            rightMargin=1.5*cm, leftMargin=1.5*cm,
            topMargin=2*cm, bottomMargin=2*cm)

        styles   = getSampleStyleSheet()
        title_st = ParagraphStyle('T', parent=styles['Title'],
            fontSize=18, spaceAfter=6, textColor=colors.HexColor('#0c1f3f'))
        sub_st   = ParagraphStyle('S', parent=styles['Normal'],
            fontSize=10, textColor=colors.grey, spaceAfter=14)

        story = []
        story.append(Paragraph("Radhika Apartment Co-op Housing Society", title_st))
        story.append(Paragraph(
            f"Maintenance Report | Generated: {datetime.date.today().strftime('%d %B %Y')} | "
            f"Wing: {wing or 'All'} | Status: {status_filter or 'All'}", sub_st))
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a56db')))
        story.append(Spacer(1, 14))

        total_balance = sum(f.total_balance for f in flats)
        summary_data  = [
            ['Total Flats', 'Paid', 'Pending', 'Overdue', 'Total Outstanding'],
            [str(len(flats)),
             str(sum(1 for f in flats if f.payment_status=='paid')),
             str(sum(1 for f in flats if f.payment_status=='pending')),
             str(sum(1 for f in flats if f.payment_status=='overdue')),
             f'Rs. {total_balance:,.0f}']
        ]
        sum_tbl = Table(summary_data, colWidths=[3.5*cm]*5)
        sum_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0),(-1,0), colors.HexColor('#0c1f3f')),
            ('TEXTCOLOR',  (0,0),(-1,0), colors.white),
            ('FONTNAME',   (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',   (0,0),(-1,-1), 9),
            ('ALIGN',      (0,0),(-1,-1), 'CENTER'),
            ('BACKGROUND', (0,1),(-1,1),  colors.HexColor('#eff6ff')),
            ('FONTNAME',   (0,1),(-1,1),  'Helvetica-Bold'),
            ('GRID',       (0,0),(-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('TOPPADDING', (0,0),(-1,-1), 6),
            ('BOTTOMPADDING',(0,0),(-1,-1),6),
        ]))
        story.append(sum_tbl)
        story.append(Spacer(1, 16))

        headers    = ['Flat No.','Owner Name','Phone','Wing','Floor','Monthly (Rs.)','Balance (Rs.)','Status','Last Payment']
        table_data = [headers]
        row_styles = []

        for i, flat in enumerate(flats, start=1):
            owner       = getattr(flat, 'owner', None)
            last_pay    = flat.payments.filter(status='completed').order_by('-payment_date').first()
            last_pay_str= last_pay.payment_date.strftime('%d/%m/%Y') if last_pay else 'Never'
            table_data.append([
                flat.flat_number,
                owner.name  if owner else 'N/A',
                owner.phone if owner else 'N/A',
                f'Wing {flat.wing.name}',
                flat.floor.name,
                f'{flat.monthly_maintenance:,.0f}',
                f'{flat.total_balance:,.0f}',
                flat.payment_status.upper(),
                last_pay_str,
            ])
            row_styles.append(('BACKGROUND',(0,i),(-1,i),
                STATUS_COLORS_PDF.get(flat.payment_status, colors.white)))

        col_w = [2.5*cm,4.5*cm,3.2*cm,2.2*cm,3*cm,3*cm,3*cm,2.5*cm,3*cm]
        tbl   = Table(table_data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#0c1f3f')),
            ('TEXTCOLOR', (0,0),(-1,0), colors.white),
            ('FONTNAME',  (0,0),(-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',  (0,0),(-1,-1), 8),
            ('ALIGN',     (0,0),(-1,-1), 'CENTER'),
            ('ALIGN',     (1,1),(2,-1),  'LEFT'),
            ('GRID',      (0,0),(-1,-1), 0.4, colors.HexColor('#e5e7eb')),
            ('TOPPADDING',(0,0),(-1,-1), 5),
            ('BOTTOMPADDING',(0,0),(-1,-1),5),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#f8fafc')]),
        ] + row_styles))
        story.append(tbl)

        doc.build(story)
        buffer.seek(0)
        resp = HttpResponse(buffer, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="maintenance_report_{datetime.date.today()}.pdf"'
        resp['Access-Control-Allow-Origin'] = '*'
        return resp


# ── INDIVIDUAL FLAT RECEIPT PDF ────────────────────────────────────────
class FlatReceiptPDFView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, flat_number):
        try:
            flat = Flat.objects.select_related('wing','floor','owner').get(flat_number=flat_number)
        except Flat.DoesNotExist:
            from rest_framework.response import Response
            from rest_framework import status
            return Response({'error':'Flat not found'}, status=status.HTTP_404_NOT_FOUND)

        owner    = getattr(flat, 'owner', None)
        payments = flat.payments.filter(status='completed').order_by('-payment_date')[:10]

        buffer = io.BytesIO()
        doc    = SimpleDocTemplate(buffer, pagesize=A4,
            rightMargin=2*cm, leftMargin=2*cm,
            topMargin=2*cm, bottomMargin=2*cm)

        styles  = getSampleStyleSheet()
        story   = []

        # ── Header ──
        header_data = [[
            Paragraph('<b><font size=14 color="#0c1f3f">Radhika Apartment</font></b><br/>'
                      '<font size=8 color="grey">Co-op Housing Society (Niyojit)</font><br/>'
                      '<font size=7 color="grey">New Ganesh Nagar, Hazimalang Road, Adiwali, Kalyan (East) - 421306</font>',
                      styles['Normal']),
            Paragraph(f'<b><font size=22 color="#1a56db">RECEIPT</font></b><br/>'
                      f'<font size=8 color="grey">Date: {datetime.date.today().strftime("%d %B %Y")}</font>',
                      ParagraphStyle('R', alignment=2, parent=styles['Normal'])),
        ]]
        header_tbl = Table(header_data, colWidths=[12*cm, 6*cm])
        header_tbl.setStyle(TableStyle([
            ('VALIGN',(0,0),(-1,-1),'TOP'),
            ('BOTTOMPADDING',(0,0),(-1,-1),12),
        ]))
        story.append(header_tbl)
        story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#1a56db')))
        story.append(Spacer(1,16))

        # ── Flat & Owner Info ──
        sc = flat.payment_status
        bal_color = '#15803d' if sc=='paid' else '#991b1b' if sc=='overdue' else '#854d0e'

        info_data = [
            ['Flat Number',  flat.flat_number,    'Wing',    f'Wing {flat.wing.name}'],
            ['Floor',        flat.floor.name,      'Status',
             Paragraph(f'<b><font color="{bal_color}">{sc.upper()}</font></b>', styles['Normal'])],
            ['Owner Name',   owner.name  if owner else 'N/A', 'Phone', owner.phone if owner else 'N/A'],
            ['Email',        owner.email if owner else 'N/A', 'Monthly', f'Rs. {flat.monthly_maintenance:,.0f}'],
        ]
        info_tbl = Table(info_data, colWidths=[4*cm,6*cm,4*cm,4*cm])
        info_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(0,-1), colors.HexColor('#f0f4f8')),
            ('BACKGROUND',(2,0),(2,-1), colors.HexColor('#f0f4f8')),
            ('FONTNAME',  (0,0),(0,-1), 'Helvetica-Bold'),
            ('FONTNAME',  (2,0),(2,-1), 'Helvetica-Bold'),
            ('FONTSIZE',  (0,0),(-1,-1), 9),
            ('GRID',      (0,0),(-1,-1), 0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING',(0,0),(-1,-1), 7),
            ('BOTTOMPADDING',(0,0),(-1,-1),7),
            ('LEFTPADDING',(0,0),(-1,-1),8),
        ]))
        story.append(info_tbl)
        story.append(Spacer(1,16))

        # ── Outstanding Balance Box ──
        bal_bg = colors.HexColor('#dcfce7') if sc=='paid' else colors.HexColor('#fee2e2') if sc=='overdue' else colors.HexColor('#fef9c3')
        bal_data = [[
            Paragraph(f'<b><font size=11>Outstanding Balance</font></b><br/>'
                      f'<font size=22 color="{bal_color}"><b>Rs. {flat.total_balance:,.0f}</b></font>',
                      styles['Normal']),
            Paragraph(f'<font size=9 color="grey">Pending Months: <b>{int(flat.total_balance/flat.monthly_maintenance) if flat.total_balance>0 else 0}</b><br/>'
                      f'Monthly Charge: <b>Rs. {flat.monthly_maintenance:,.0f}</b></font>',
                      styles['Normal']),
        ]]
        bal_tbl = Table(bal_data, colWidths=[9*cm,9*cm])
        bal_tbl.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,-1), bal_bg),
            ('GRID',(0,0),(-1,-1),0.5, colors.HexColor('#e2e8f0')),
            ('TOPPADDING',(0,0),(-1,-1),12),
            ('BOTTOMPADDING',(0,0),(-1,-1),12),
            ('LEFTPADDING',(0,0),(-1,-1),14),
            ('ROUNDEDCORNERS',[6]),
        ]))
        story.append(bal_tbl)
        story.append(Spacer(1,20))

        # ── Payment History ──
        story.append(Paragraph('<b><font size=11 color="#0c1f3f">Recent Payment History</font></b>',
                                styles['Normal']))
        story.append(Spacer(1,8))

        if payments:
            pay_headers = ['#','Date','Amount (Rs.)','Mode','Transaction ID','Received By']
            pay_data    = [pay_headers]
            for i, pay in enumerate(payments,1):
                pay_data.append([
                    str(i),
                    pay.payment_date.strftime('%d/%m/%Y'),
                    f'{pay.amount:,.0f}',
                    pay.payment_mode.upper(),
                    pay.transaction_id or '—',
                    pay.received_by.get_full_name() if pay.received_by else 'Admin',
                ])
            pay_tbl = Table(pay_data, colWidths=[1*cm,3.5*cm,4*cm,3*cm,4*cm,3*cm])
            pay_tbl.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0), colors.HexColor('#0c1f3f')),
                ('TEXTCOLOR', (0,0),(-1,0), colors.white),
                ('FONTNAME',  (0,0),(-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',  (0,0),(-1,-1), 8),
                ('ALIGN',     (0,0),(-1,-1), 'CENTER'),
                ('GRID',      (0,0),(-1,-1), 0.4, colors.HexColor('#e5e7eb')),
                ('TOPPADDING',(0,0),(-1,-1), 5),
                ('BOTTOMPADDING',(0,0),(-1,-1),5),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#f0f9ff')]),
            ]))
            story.append(pay_tbl)
        else:
            story.append(Paragraph('<font color="grey">No payment records found.</font>',
                                    styles['Normal']))

        story.append(Spacer(1,30))
        story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#e2e8f0')))
        story.append(Spacer(1,8))
        story.append(Paragraph(
            '<font size=8 color="grey">This is a computer-generated receipt. '
            'For queries contact the society office.</font>',
            ParagraphStyle('Footer', alignment=1, parent=styles['Normal'])))

        doc.build(story)
        buffer.seek(0)
        resp = HttpResponse(buffer, content_type='application/pdf')
        resp['Content-Disposition'] = f'attachment; filename="receipt_{flat_number}_{datetime.date.today()}.pdf"'
        resp['Access-Control-Allow-Origin'] = '*'
        return resp


# ── EXCEL EXPORT ───────────────────────────────────────────────────────
class ExportExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        flats         = get_flat_data()
        wing          = request.query_params.get('wing')
        status_filter = request.query_params.get('status')
        if wing:
            flats = [f for f in flats if f.wing.name == wing.upper()]
        if status_filter:
            flats = [f for f in flats if f.payment_status == status_filter]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maintenance Summary"

        hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        hdr_fill  = PatternFill(start_color='0c1f3f', end_color='0c1f3f', fill_type='solid')
        center    = Alignment(horizontal='center', vertical='center')
        thin      = Side(style='thin', color='D1D5DB')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        STATUS_FILLS = {
            'paid':    PatternFill(start_color='DCFCE7', end_color='DCFCE7', fill_type='solid'),
            'pending': PatternFill(start_color='FEF9C3', end_color='FEF9C3', fill_type='solid'),
            'overdue': PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid'),
        }

        ws.merge_cells('A1:I1')
        ws['A1'] = 'Radhika Apartment Co-op Housing Society — Maintenance Report'
        ws['A1'].font      = Font(name='Calibri', bold=True, size=14, color='0c1f3f')
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 28

        month = request.query_params.get('month')
        year  = request.query_params.get('year')

        ws.merge_cells('A2:I2')
        summary_parts = [f"Generated: {datetime.date.today().strftime('%d %B %Y')}",
                          f"Wing: {wing or 'All'}", f"Status: {status_filter or 'All'}"]
        if month or year:
            MONTH_NAMES = ['','January','February','March','April','May','June',
                           'July','August','September','October','November','December']
            label = (MONTH_NAMES[int(month)] if month else '') + (f' {year}' if year else '')
            summary_parts.append(f"Payments Month: {label.strip()}")
        ws['A2'] = " | ".join(summary_parts)
        ws['A2'].font      = Font(name='Calibri', size=10, italic=True, color='6B7280')
        ws['A2'].alignment = center

        headers = ['Flat No.','Owner Name','Phone','Email','Wing','Floor','Monthly (Rs.)','Balance (Rs.)','Status']
        for col, h in enumerate(headers, 1):
            cell            = ws.cell(row=4, column=col, value=h)
            cell.font       = hdr_font
            cell.fill       = hdr_fill
            cell.alignment  = center
            cell.border     = border

        ws.row_dimensions[4].height = 22
        for i, w in enumerate([12,22,15,28,8,14,16,16,12], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        for row_idx, flat in enumerate(flats, start=5):
            owner = getattr(flat, 'owner', None)
            data  = [
                flat.flat_number,
                owner.name  if owner else 'N/A',
                owner.phone if owner else 'N/A',
                owner.email if owner else 'N/A',
                f'Wing {flat.wing.name}',
                flat.floor.name,
                float(flat.monthly_maintenance),
                float(flat.total_balance),
                flat.payment_status.upper(),
            ]
            fill = STATUS_FILLS.get(flat.payment_status)
            for col_idx, val in enumerate(data, 1):
                cell            = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border     = border
                cell.alignment  = center
                if fill: cell.fill = fill
                if col_idx in [7,8]: cell.number_format = '#,##0.00'

        # ── Sheet 2: Payment History (filterable by month/year) ──
        payments_qs = Payment.objects.select_related('flat','flat__owner','received_by').order_by('-payment_date')
        if month:
            payments_qs = payments_qs.filter(payment_date__month=int(month))
        if year:
            payments_qs = payments_qs.filter(payment_date__year=int(year))

        sheet2_title = "Payment History"
        if month or year:
            MONTH_NAMES = ['','January','February','March','April','May','June',
                           'July','August','September','October','November','December']
            label = ''
            if month: label += MONTH_NAMES[int(month)]
            if year:  label += f' {year}'
            sheet2_title = f"Payments - {label.strip()}"

        ws2 = wb.create_sheet(sheet2_title[:31])  # Excel sheet name max 31 chars

        # Title row for sheet 2
        ws2.merge_cells('A1:H1')
        title2 = "Payment History"
        if month or year:
            title2 += f" — {sheet2_title.replace('Payments - ','')}"
        ws2['A1'] = title2
        ws2['A1'].font      = Font(name='Calibri', bold=True, size=13, color='0c1f3f')
        ws2['A1'].alignment = center
        ws2.row_dimensions[1].height = 24

        header_row = 2
        for col, h in enumerate(['Flat No.','Owner','Amount (Rs.)','Date','Mode','Transaction ID','Status','Received By'], 1):
            cell = ws2.cell(row=header_row, column=col, value=h)
            cell.font=hdr_font; cell.fill=hdr_fill; cell.alignment=center; cell.border=border

        total_collected = 0
        row_idx = header_row + 1
        for pay in payments_qs[:1000]:
            owner = getattr(pay.flat,'owner',None)
            for col_idx, val in enumerate([
                pay.flat.flat_number,
                owner.name if owner else 'N/A',
                float(pay.amount),
                str(pay.payment_date),
                pay.payment_mode,
                pay.transaction_id or '',
                pay.status.upper(),
                pay.received_by.get_full_name() if pay.received_by else '',
            ], 1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = center
                if col_idx == 3: cell.number_format = '#,##0.00'
            if pay.status == 'completed':
                total_collected += float(pay.amount)
            row_idx += 1

        # Total row
        ws2.cell(row=row_idx, column=2, value='TOTAL COLLECTED').font = Font(bold=True)
        ws2.cell(row=row_idx, column=2).alignment = center
        ws2.cell(row=row_idx, column=2).border = border
        total_cell = ws2.cell(row=row_idx, column=3, value=total_collected)
        total_cell.font = Font(bold=True, color='15803d')
        total_cell.number_format = '#,##0.00'
        total_cell.alignment = center
        total_cell.border = border
        ws2.cell(row=row_idx, column=1).border = border

        for i in range(1,9):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

        output = io.BytesIO()
        wb.save(output); output.seek(0)
        resp = HttpResponse(output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename_suffix = datetime.date.today().isoformat()
        if month or year:
            filename_suffix = f"{year or datetime.date.today().year}-{(month or '').zfill(2) if month else 'all'}"
        resp['Content-Disposition'] = f'attachment; filename="maintenance_{filename_suffix}.xlsx"'
        resp['Access-Control-Allow-Origin'] = '*'
        return resp
